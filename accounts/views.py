from django.shortcuts import render, Http404
from django.template import RequestContext
from django.http import HttpResponse,  HttpResponseRedirect, JsonResponse
from django.contrib.auth.decorators import login_required
import urllib, json
import time
from datetime import date, timedelta
import math
import random
from urllib.parse import urlparse
from os.path import splitext
from .utils import generate_token, get_query, normalize_query
# from django.contrib.gis.utils import GeoIP
from django.core.mail import send_mail, EmailMessage
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.contrib import messages
from django.db.models import Count, Min, Sum, Avg, Max, Q
from django.conf import settings
from django.urls import reverse
from .models import *
from .forms import *
from nominations.models import *
from nominations.forms import *
from votes.models import *
# from ideas.models import *
# from ideas.forms import *
# from payments.models import *
from blog.models import *
from blog.forms import *
from django.views import *
from django.core.paginator import *
from vendor.functions.tasks import *

# Openpyxl
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

class AccountView(View):
    template_account = 'account.html'
        
    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        profile = Accounts.objects.filter(user=request.user).last() 
        return HttpResponseRedirect(reverse('superadmin_schemes'))
    
    """This is responsible for detecting post method and processing information """
    def post(self, request):
        return HttpResponseRedirect(reverse('superadmin_schemes'))

class AccountSchemesView(View):
    template_accountpanel = 'accountpanel.html'
        
    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        profile = Accounts.objects.filter(user=request.user).last()
        allschemes = AwardSchemes.objects.filter()
        paginator = Paginator(allschemes, 100) # Show 100 contacts per page		
        page = request.GET.get('page')
        schemes = paginator.get_page(page)	
        return render(request, self.template_accountpanel, { 'accountinfo':profile, 'schemes':schemes    })
    
    """This is responsible for detecting post method and processing information """
    def post(self, request):
        return HttpResponseRedirect(reverse('superadmin_schemes'))

class AccountSchemesExcelView(View):
    form_class = ''
    initial = {'': ''}
    template_accountpanel = 'accountpanel.html'

    """This is responsible for detecting get method and redirecting """
    def get(self, request):
        profile = Accounts.objects.filter(user=request.user).last()        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Schemes Reports"
            ws.subject = "Schemes Reports"
            ws.description = "This is a spreadsheet of " + "Schemes Reports"
            ws['A1'] = "Humanitarian Awards Global"
            ws.merge_cells('A1:H1')
            ws['A2'] = 'HAG'
            ws.merge_cells('A2:H2')
            ws['A4'] = "Schemes Reports"


            ws['A7'] = 'Year'				
            ws['B7'] = 'Name'				
            ws['C7'] = 'Theme'
            ws['D7'] = 'Description'	
            ws['E7'] = 'Active'	

            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 60
            ws.column_dimensions['E'].width = 10

            schemes = AwardSchemes.objects.filter().values_list('id', 'year', 'name', 'theme', 'description', 'active')
            ite = 1
            for scheme in schemes:
                ws['A'+str(ite + 7)] = str(scheme[1] or '')
                ws['B'+str(ite + 7)] = str(scheme[2] or '')
                ws['C'+str(ite + 7)] = str(scheme[3] or '')
                ws['D'+str(ite + 7)] = str(scheme[4] or '')
                ws['E'+str(ite + 7)] = str(scheme[5] or '')

                ite += 1
      
            # useractivity = UserActivity(school=profile.school, activity=profile.fullname + ' generated excel reports for ' + str(reporttype or '') +' at ' +str(datetime.now()))
            # useractivity.save()

            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
            ws.page_setup.fitToHeight = 0
            ws.page_setup.fitToWidth = 1
            output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            file_name = "Schemes_Reports.xlsx"
            output['Content-Disposition'] = 'attachment; filename='+ file_name
            wb.save(output)
            return output
            
        except Exception as e:
            return HttpResponse(str(e))

        return render(request, self.template_accountpanel, { 
                                                        'accountinfo': profile })

    def post(self, request):
        profile = UserProfile.objects.filter(user=request.user).last()
        if not profile.module_reports:
            return HttpResponseRedirect(reverse('account'))
            
        return render(request, self.template_accountpanel, { 
                                                        'accountinfo': profile })

class AccountCategoriesView(View):
    template_accountpanel = 'accountpanel.html'
        
    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        profile = Accounts.objects.filter(user=request.user).last()
        allcategories = NominationCategories.objects.filter().order_by('-year', 'name')
        paginator = Paginator(allcategories, 100) # Show 100 contacts per page		
        page = request.GET.get('page')
        categories = paginator.get_page(page)	
        return render(request, self.template_accountpanel, { 'accountinfo':profile, 'categories':categories    })
    
    """This is responsible for detecting post method and processing information """
    def post(self, request):
        return HttpResponseRedirect(reverse('superadmin_categories'))

class AccountCategoriesExcelView(View):
    form_class = ''
    initial = {'': ''}
    template_accountpanel = 'accountpanel.html'

    """This is responsible for detecting get method and redirecting """
    def get(self, request):
        profile = Accounts.objects.filter(user=request.user).last()        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Nomination Categories Reports"
            ws.subject = "Nomination Categories Reports"
            ws.description = "This is a spreadsheet of " + "Nomination Categories Reports"
            ws['A1'] = "Humanitarian Awards Global"
            ws.merge_cells('A1:H1')
            ws['A2'] = 'HAG'
            ws.merge_cells('A2:H2')
            ws['A4'] = "Nomination Categories Reports"


            ws['A7'] = 'Scheme'				
            ws['B7'] = 'Year'				
            ws['C7'] = 'Name'
            ws['D7'] = 'Description'	
            ws['E7'] = 'Active'	

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 60
            ws.column_dimensions['E'].width = 10

            categories = NominationCategories.objects.filter().values_list('id', 'scheme__year', 'scheme__name', 'year', 'name', 'description_long', 'active')
            ite = 1
            for category in categories:
                ws['A'+str(ite + 7)] = str(category[1] or '')
                ws['B'+str(ite + 7)] = str(category[2] or '')
                ws['C'+str(ite + 7)] = str(category[3] or '')
                ws['D'+str(ite + 7)] = str(category[4] or '')
                ws['E'+str(ite + 7)] = str(category[5] or '')

                ite += 1
      
            # useractivity = UserActivity(school=profile.school, activity=profile.fullname + ' generated excel reports for ' + str(reporttype or '') +' at ' +str(datetime.now()))
            # useractivity.save()

            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
            ws.page_setup.fitToHeight = 0
            ws.page_setup.fitToWidth = 1
            output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            file_name = "Nomination_Categories_Reports.xlsx"
            output['Content-Disposition'] = 'attachment; filename='+ file_name
            wb.save(output)
            return output
            
        except Exception as e:
            return HttpResponse(str(e))

        return render(request, self.template_accountpanel, { 
                                                        'accountinfo': profile })

    def post(self, request):
        profile = UserProfile.objects.filter(user=request.user).last()
        if not profile.module_reports:
            return HttpResponseRedirect(reverse('account'))
            
        return render(request, self.template_accountpanel, { 
                                                        'accountinfo': profile })

class AccountNominationsView(View):
    form_class = NomineeForm
    initial = {'': ''}
    template_accountpanel = 'accountpanel.html'
        
    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        profile = Accounts.objects.filter(user=request.user).last()
        scheme_get = request.GET.get('scheme')
        category_get = request.GET.get('category')
        searchvalue = request.GET.get('q')
        nomineeid = request.GET.get('nomineeid')
        Database_Filters = {  }

        schemes = AwardSchemes.objects.filter().order_by('year')
        scheme = AwardSchemes.objects.filter(id=scheme_get).last()
        if scheme:
            Database_Filters['category__scheme__id'] = scheme.id
        else:
            scheme = AwardSchemes.objects.filter(latest=True).order_by('-id').last()
            if scheme:
                Database_Filters['category__scheme__id'] = scheme.id

        categories = NominationCategories.objects.filter(scheme=scheme).order_by('-year', 'name')
        category = NominationCategories.objects.filter(scheme__id=scheme.id, id=category_get).last()
        if category:
            Database_Filters['category__id'] = category.id
        
        if searchvalue:
            allnominations = Nominations.objects.filter(Q(referencenumber__contains=searchvalue) | Q(nominee_firstname__icontains=searchvalue) | Q(nominee_middlename__icontains=searchvalue) | Q(nominee_lastname__icontains=searchvalue) | Q(nominee_phonenumber__contains=searchvalue) | Q(nominee_emailaddress__icontains=searchvalue) | Q(nominee_agegroup__icontains=searchvalue) | Q(nominee_country__icontains=searchvalue) | Q(nominee_code__contains=searchvalue), **Database_Filters).order_by('-approved', '-shortlisted', 'category__year', 'category__name')
        else:
            allnominations = Nominations.objects.filter(**Database_Filters).order_by('-approved', '-shortlisted', 'category__year', 'category__name')
        

        # allnominations = Nominations.objects.filter(**Database_Filters).order_by('id')
        paginator = Paginator(allnominations, 100) # Show 100 contacts per page		
        page = request.GET.get('page')
        nominations = paginator.get_page(page)	

        nominee = Nominations.objects.filter(id=nomineeid).last()
        if nominee:
            form = self.form_class(instance=nominee)
        else:
            form = self.form_class(initial=self.initial)

        return render(request, self.template_accountpanel, { 'accountinfo':profile, 'schemes':schemes, 'scheme':scheme, 'categories':categories, 'category':category, 'q':searchvalue, 'nominations':nominations, 'nominee':nominee, 'form':form,   })
    
    """This is responsible for detecting post method and processing information """
    def post(self, request):
        if 'submit' in request.POST:            
            return HttpResponseRedirect(build_url('superadmin_nominations', kwargs={}, params={}))
        elif 'update' in request.POST:
            nominee = Nominations.objects.filter(id=request.POST.get('id')).last()
            form = self.form_class(request.POST, request.FILES, instance=nominee)
            if form.is_valid():
                instance = form.save(commit=False)
                instance.save()
                messages.success(request, 'Nominee updated Successfully')
            else:
                messages.error(request, 'Nominee could not be updated due to ' + str(form.errors or ''))

            return HttpResponseRedirect(build_url('superadmin_nominations', kwargs={}, params={'nomineeid': nominee.id}))

        return HttpResponseRedirect(build_url('superadmin_nominations', kwargs={}, params={}))

        return HttpResponseRedirect(reverse('superadmin_nominations'))

class AccountNominationsExcelView(View):
    form_class = ''
    initial = {'': ''}
    template_accountpanel = 'accountpanel.html'

    """This is responsible for detecting get method and redirecting """
    def get(self, request):
        report_type = request.GET.get('report_type')
        # return HttpResponse(report_type)
        profile = Accounts.objects.filter(user=request.user).last()        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Nominations Reports"
            ws.subject = "Nominations Reports"
            ws.description = "This is a spreadsheet of " + "Nominations Reports"
            ws['A1'] = "Humanitarian Awards Global"
            ws.merge_cells('A1:H1')
            ws['A2'] = 'HAG'
            ws.merge_cells('A2:H2')
            ws['A4'] = "Nominations Reports"


            ws['A7'] = 'Category'				
            ws['B7'] = 'Reference No'				
            ws['C7'] = 'Nominee Firstname'
            ws['D7'] = 'Nominee Middlename'	
            ws['E7'] = 'Nominee Lastname'	
            ws['F7'] = 'Nominee Phonenumber'	
            ws['G7'] = 'Nominee Email Address'	
            ws['H7'] = 'Gender'	
            ws['I7'] = 'Age Group'	
            ws['J7'] = 'Nominee Bio'
            ws['K7'] = 'Why Nominee Deserves It'	
            ws['L7'] = 'Efforts by Nominee'	
            ws['M7'] = 'Turning Point by Nominee'	
            ws['N7'] = 'What makes Nominee Exceptional'	
            ws['O7'] = 'Nominee Reference Links'	
            ws['P7'] = 'Nominee Town or City'
            ws['Q7'] = 'Nominee Region'	
            ws['R7'] = 'Nominee Country'
            ws['S7'] = 'Nominee Code'
            ws['T7'] = 'Company Name'
            ws['U7'] = 'Company Industry'
            ws['V7'] = 'Company Bio'
            ws['W7'] = 'Nominator Fullname'
            ws['X7'] = 'Nominator Phonenumber'
            ws['Y7'] = 'Nominator Email Address'
            ws['Z7'] = 'Nominator Bio'
            ws['AA7'] = 'Time Known Nominee'
            ws['AB7'] = 'Relationship with Nominee'
            ws['AC7'] = 'Shortlisted'
            ws['AD7'] = 'Approved'
            ws['AE7'] = 'Votable'

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 70
            ws.column_dimensions['K'].width = 70
            ws.column_dimensions['L'].width = 70
            ws.column_dimensions['M'].width = 70
            ws.column_dimensions['N'].width = 70
            ws.column_dimensions['O'].width = 20
            ws.column_dimensions['P'].width = 20
            ws.column_dimensions['Q'].width = 20
            ws.column_dimensions['R'].width = 20
            ws.column_dimensions['S'].width = 20
            ws.column_dimensions['T'].width = 30
            ws.column_dimensions['U'].width = 20
            ws.column_dimensions['V'].width = 20
            ws.column_dimensions['W'].width = 20
            ws.column_dimensions['X'].width = 20
            ws.column_dimensions['Y'].width = 20
            ws.column_dimensions['Z'].width = 20
            ws.column_dimensions['AA'].width = 20
            ws.column_dimensions['AB'].width = 20
            ws.column_dimensions['AC'].width = 20
            ws.column_dimensions['AD'].width = 20
            ws.column_dimensions['AE'].width = 20

            allnominations = None
            
            if report_type == 'All Nominations':
                allnominations = Nominations.objects.filter()
            elif report_type == 'All Shortlisted Nominations':
                allnominations = Nominations.objects.filter(shortlisted=True)
            elif report_type == '2020 - Shortlisted':
                allnominations = Nominations.objects.filter(category__scheme__year='2020', shortlisted=True)
            elif report_type == '2021 - Shortlisted':
                allnominations = Nominations.objects.filter(category__scheme__year='2021', shortlisted=True)
            elif report_type == '2022 - Shortlisted':
                allnominations = Nominations.objects.filter(category__scheme__year='2022', shortlisted=True)
            elif report_type == '2023 - Shortlisted':
                allnominations = Nominations.objects.filter(category__scheme__year='2023', shortlisted=True)
            elif report_type == '2024 - Shortlisted':
                allnominations = Nominations.objects.filter(category__scheme__year='2024', shortlisted=True)
            elif report_type == '2025 - Shortlisted':
                allnominations = Nominations.objects.filter(category__scheme__year='2025', shortlisted=True)
            elif report_type == '2026 - Shortlisted':
                allnominations = Nominations.objects.filter(category__scheme__year='2026', shortlisted=True)
            elif report_type == '2027 - Shortlisted':
                allnominations = Nominations.objects.filter(category__scheme__year='2027', shortlisted=True)
            else:
                allnominations = Nominations.objects.filter()
            
            nominations = allnominations.order_by('id').values_list('id', 'category__year', 'category__name', 'referencenumber', 'nominee_firstname', 'nominee_middlename', 'nominee_lastname', 'nominee_phonenumber', 'nominee_emailaddress', 'nominee_gender', 'nominee_agegroup', 'nominee_bio', 'nominee_deserve_info', 'nominee_effort_info', 'nominee_turning_point', 'nominee_exception', 'nominee_reference_links', 'nominee_towncity', 'nominee_region', 'nominee_country', 'nominee_code', 'company_name', 'company_industry', 'company_bio', 'nominator_fullname', 'nominator_phonenumber', 'nominator_emailaddress', 'nominator_bio', 'nominator_knownnominee_time', 'nominator_nominee_relationship', 'shortlisted', 'approved', 'votable')
            
            ite = 1
            
            for nomination in nominations:
                ws['A'+str(ite + 7)] = str(nomination[1] or '') + ' - ' + str(nomination[2] or '')
                ws['B'+str(ite + 7)] = str(nomination[3] or '')
                ws['C'+str(ite + 7)] = str(nomination[4] or '')
                ws['D'+str(ite + 7)] = str(nomination[5] or '')
                ws['E'+str(ite + 7)] = str(nomination[6] or '')
                ws['F'+str(ite + 7)] = str(nomination[7] or '')
                ws['G'+str(ite + 7)] = str(nomination[8] or '')
                ws['H'+str(ite + 7)] = str(nomination[9] or '')
                ws['I'+str(ite + 7)] = str(nomination[10] or '')
                ws['J'+str(ite + 7)] = str(nomination[11] or '')
                ws['K'+str(ite + 7)] = str(nomination[12] or '')
                ws['L'+str(ite + 7)] = str(nomination[13] or '')
                ws['M'+str(ite + 7)] = str(nomination[14] or '')
                ws['N'+str(ite + 7)] = str(nomination[15] or '')
                ws['O'+str(ite + 7)] = str(nomination[16] or '')
                ws['P'+str(ite + 7)] = str(nomination[17] or '')
                ws['Q'+str(ite + 7)] = str(nomination[18] or '')
                ws['R'+str(ite + 7)] = str(nomination[19] or '')
                ws['S'+str(ite + 7)] = str(nomination[20] or '')
                ws['T'+str(ite + 7)] = str(nomination[21] or '')
                ws['U'+str(ite + 7)] = str(nomination[22] or '')
                ws['V'+str(ite + 7)] = str(nomination[23] or '')
                ws['W'+str(ite + 7)] = str(nomination[24] or '')
                ws['X'+str(ite + 7)] = str(nomination[25] or '')
                ws['Y'+str(ite + 7)] = str(nomination[26] or '')
                ws['Z'+str(ite + 7)] = str(nomination[27] or '')
                ws['AA'+str(ite + 7)] = str(nomination[28] or '')
                ws['AB'+str(ite + 7)] = str(nomination[29] or '')
                ws['AC'+str(ite + 7)] = str(nomination[30] or '')
                ws['AD'+str(ite + 7)] = str(nomination[31] or '')
                ws['AE'+str(ite + 7)] = str(nomination[32] or '')

                ite += 1
      
            # useractivity = UserActivity(school=profile.school, activity=profile.fullname + ' generated excel reports for ' + str(reporttype or '') +' at ' +str(datetime.now()))
            # useractivity.save()

            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
            ws.page_setup.fitToHeight = 0
            ws.page_setup.fitToWidth = 1
            output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            file_name = "Nominations_Reports.xlsx"
            output['Content-Disposition'] = 'attachment; filename='+ file_name
            wb.save(output)
            return output
            
        except Exception as e:
            return HttpResponse(str(e))

        return render(request, self.template_accountpanel, { 
                                                        'accountinfo': profile })

    def post(self, request):
        profile = UserProfile.objects.filter(user=request.user).last()
        if not profile.module_reports:
            return HttpResponseRedirect(reverse('account'))
            
        return render(request, self.template_accountpanel, { 
                                                        'accountinfo': profile })

class AccountVotesView(View):
    form_class = NomineeForm
    initial = {'': ''}
    template_accountpanel = 'accountpanel.html'
        
    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        profile = Accounts.objects.filter(user=request.user).last()
        scheme_get = request.GET.get('scheme')
        category_get = request.GET.get('category')
        searchvalue = request.GET.get('q')
        nomineeid = request.GET.get('nomineeid')
        Database_Filters = {  }

        schemes = AwardSchemes.objects.filter().order_by('year')
        scheme = AwardSchemes.objects.filter(id=scheme_get).last()
        if scheme:
            Database_Filters['category__scheme__id'] = scheme.id
        else:
            scheme = AwardSchemes.objects.filter(latest=True).order_by('-id').last()
            if scheme:
                Database_Filters['category__scheme__id'] = scheme.id

        categories = NominationCategories.objects.filter(scheme=scheme).order_by('-year', 'name')
        category = NominationCategories.objects.filter(scheme__id=scheme.id, id=category_get).last()
        if category:
            Database_Filters['category__id'] = category.id
        
        if searchvalue:
            allnominations = Nominations.objects.filter(Q(referencenumber__contains=searchvalue) | Q(nominee_firstname__icontains=searchvalue) | Q(nominee_middlename__icontains=searchvalue) | Q(nominee_lastname__icontains=searchvalue) | Q(nominee_phonenumber__contains=searchvalue) | Q(nominee_emailaddress__icontains=searchvalue) | Q(nominee_agegroup__icontains=searchvalue) | Q(nominee_country__icontains=searchvalue) | Q(nominee_code__contains=searchvalue), approved=True, **Database_Filters).order_by('id', 'category__year', 'category__name', 'approved')
        else:
            allnominations = Nominations.objects.filter(approved=True, **Database_Filters).order_by('id', 'category__year', 'category__name', 'approved')
        
        allvotes = []
        for nominee in allnominations:
            votecount_chk = Votes.objects.filter(nomination=nominee).aggregate(Sum('votecount'))
            votecount = str(float(votecount_chk['votecount__sum'] or 0) or '0')
            d = {'id':nominee.id, 'fullname':str(nominee.nominee_firstname or '')+' '+str(nominee.nominee_lastname or ''), 'scheme':nominee.category.scheme.id, 'category':nominee.category.id, 'code':nominee.nominee_code, 'votes':votecount }
            allvotes.append(d)

        # allnominations = Votes.objects.filter(**Database_Filters).order_by('id')
        paginator = Paginator(allvotes, 100) # Show 100 contacts per page		
        page = request.GET.get('page')
        votes = paginator.get_page(page)	

        return render(request, self.template_accountpanel, { 'accountinfo':profile, 'schemes':schemes, 'scheme':scheme, 'categories':categories, 'category':category, 'q':searchvalue, 'votes':votes  })
    
    """This is responsible for detecting post method and processing information """
    def post(self, request):
        return HttpResponseRedirect(build_url('superadmin_votes', kwargs={}, params={}))
        
class AccountVotesExcelView(View):
    form_class = ''
    initial = {'': ''}
    template_accountpanel = 'accountpanel.html'

    """This is responsible for detecting get method and redirecting """
    def get(self, request):
        profile = Accounts.objects.filter(user=request.user).last()        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Votes Reports"
            ws.subject = "Votes Reports"
            ws.description = "This is a spreadsheet of " + "Votes Reports"
            ws['A1'] = "Humanitarian Awards Global"
            ws.merge_cells('A1:H1')
            ws['A2'] = 'HAG'
            ws.merge_cells('A2:H2')
            ws['A4'] = "Votes Reports"


            ws['A7'] = 'Category'				
            ws['B7'] = 'Reference No'				
            ws['C7'] = 'Nominee Firstname'
            ws['D7'] = 'Nominee Middlename'	
            ws['E7'] = 'Nominee Lastname'	
            ws['F7'] = 'Nominee Phonenumber'	
            ws['G7'] = 'Nominee Email Address'	
            ws['H7'] = 'Gender'	
            ws['I7'] = 'Age Group'	
            ws['J7'] = 'Nominee Bio'
            ws['K7'] = 'Why Nominee Deserves It'	
            ws['L7'] = 'Efforts by Nominee'	
            ws['M7'] = 'Turning Point by Nominee'	
            ws['N7'] = 'What makes Nominee Exceptional'	
            ws['O7'] = 'Nominee Reference Links'	
            ws['P7'] = 'Nominee Town or City'
            ws['Q7'] = 'Nominee Region'	
            ws['R7'] = 'Nominee Country'
            ws['S7'] = 'Nominee Code'
            ws['T7'] = 'Company Name'
            ws['U7'] = 'Company Industry'
            ws['V7'] = 'Company Bio'
            ws['W7'] = 'Nominator Fullname'
            ws['X7'] = 'Nominator Phonenumber'
            ws['Y7'] = 'Nominator Email Address'
            ws['Z7'] = 'Nominator Bio'
            ws['Z7'] = 'Time Known Nominee'
            ws['Z7'] = 'Relationship with Nominee'

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 70
            ws.column_dimensions['K'].width = 70
            ws.column_dimensions['L'].width = 70
            ws.column_dimensions['M'].width = 70
            ws.column_dimensions['N'].width = 70
            ws.column_dimensions['O'].width = 20
            ws.column_dimensions['P'].width = 20
            ws.column_dimensions['Q'].width = 20
            ws.column_dimensions['R'].width = 20
            ws.column_dimensions['S'].width = 20
            ws.column_dimensions['T'].width = 30
            ws.column_dimensions['U'].width = 20
            ws.column_dimensions['V'].width = 20
            ws.column_dimensions['W'].width = 20
            ws.column_dimensions['X'].width = 20
            ws.column_dimensions['Y'].width = 20
            ws.column_dimensions['Z'].width = 20
            ws.column_dimensions['AA'].width = 20
            ws.column_dimensions['AB'].width = 20

            nominations = Votes.objects.filter().order_by('id').values_list('id', 'category__year', 'category__name', 'referencenumber', 'nominee_firstname', 'nominee_middlename', 'nominee_lastname', 'nominee_phonenumber', 'nominee_emailaddress', 'nominee_gender', 'nominee_agegroup', 'nominee_bio', 'nominee_deserve_info', 'nominee_effort_info', 'nominee_turning_point', 'nominee_exception', 'nominee_reference_links', 'nominee_towncity', 'nominee_region', 'nominee_country', 'nominee_code', 'company_name', 'company_industry', 'company_bio', 'nominator_fullname', 'nominator_phonenumber', 'nominator_emailaddress', 'nominator_bio', 'nominator_knownnominee_time', 'nominator_nominee_relationship')
            ite = 1
            for nomination in nominations:
                ws['A'+str(ite + 7)] = str(nomination[1] or '')
                ws['B'+str(ite + 7)] = str(nomination[2] or '')
                ws['C'+str(ite + 7)] = str(nomination[3] or '')
                ws['D'+str(ite + 7)] = str(nomination[4] or '')
                ws['E'+str(ite + 7)] = str(nomination[5] or '')
                ws['F'+str(ite + 7)] = str(nomination[6] or '')
                ws['G'+str(ite + 7)] = str(nomination[7] or '')
                ws['H'+str(ite + 7)] = str(nomination[8] or '')
                ws['I'+str(ite + 7)] = str(nomination[9] or '')
                ws['J'+str(ite + 7)] = str(nomination[10] or '')
                ws['K'+str(ite + 7)] = str(nomination[11] or '')
                ws['L'+str(ite + 7)] = str(nomination[12] or '')
                ws['M'+str(ite + 7)] = str(nomination[13] or '')
                ws['N'+str(ite + 7)] = str(nomination[14] or '')
                ws['O'+str(ite + 7)] = str(nomination[15] or '')
                ws['P'+str(ite + 7)] = str(nomination[16] or '')
                ws['Q'+str(ite + 7)] = str(nomination[17] or '')
                ws['R'+str(ite + 7)] = str(nomination[18] or '')
                ws['S'+str(ite + 7)] = str(nomination[19] or '')
                ws['T'+str(ite + 7)] = str(nomination[20] or '')
                ws['U'+str(ite + 7)] = str(nomination[21] or '')
                ws['V'+str(ite + 7)] = str(nomination[22] or '')
                ws['W'+str(ite + 7)] = str(nomination[23] or '')
                ws['X'+str(ite + 7)] = str(nomination[24] or '')
                ws['Y'+str(ite + 7)] = str(nomination[25] or '')
                ws['Z'+str(ite + 7)] = str(nomination[26] or '')
                ws['AA'+str(ite + 7)] = str(nomination[27] or '')
                ws['AB'+str(ite + 7)] = str(nomination[28] or '')

                ite += 1
      
            # useractivity = UserActivity(school=profile.school, activity=profile.fullname + ' generated excel reports for ' + str(reporttype or '') +' at ' +str(datetime.now()))
            # useractivity.save()

            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
            ws.page_setup.fitToHeight = 0
            ws.page_setup.fitToWidth = 1
            output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            file_name = "Votes_Reports.xlsx"
            output['Content-Disposition'] = 'attachment; filename='+ file_name
            wb.save(output)
            return output
            
        except Exception as e:
            return HttpResponse(str(e))

        return render(request, self.template_accountpanel, { 
                                                        'accountinfo': profile })

    def post(self, request):
        profile = UserProfile.objects.filter(user=request.user).last()
        if not profile.module_reports:
            return HttpResponseRedirect(reverse('account'))
            
        return render(request, self.template_accountpanel, { 
                                                        'accountinfo': profile })

class AccountPaymentsView(View):
    form_class = NomineeForm
    initial = {'': ''}
    template_accountpanel = 'accountpanel.html'
        
    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        profile = Accounts.objects.filter(user=request.user).last()
        scheme_get = request.GET.get('scheme')
        category_get = request.GET.get('category')
        searchvalue = request.GET.get('q')
        nomineeid = request.GET.get('nomineeid')
        Database_Filters = {  }

        schemes = AwardSchemes.objects.filter().order_by('year')
        scheme = AwardSchemes.objects.filter(id=scheme_get).last()
        if scheme:
            Database_Filters['nomination__category__scheme__id'] = scheme.id
        else:
            scheme = AwardSchemes.objects.filter(latest=True).order_by('-id').last()
            if scheme:
                Database_Filters['nomination__category__scheme__id'] = scheme.id
    
        if not scheme:
            messages.error(request, 'No Awards Scheme Selected')
            return HttpResponseRedirect(build_url('superadmin_payments', kwargs={}, params={}))

        categories = NominationCategories.objects.filter(scheme=scheme).order_by('-year', 'name')
        category = NominationCategories.objects.filter(scheme__id=scheme.id, id=category_get).last()
        
        if category:
            Database_Filters['nomination__category__id'] = category.id
            
        nominees = Nominations.objects.filter(category__scheme=scheme).order_by('id', 'category__year', 'category__name')
        if category and nomineeid:
            nominee = Nominations.objects.filter(category__scheme__id=scheme.id, category__id=category.id, id=nomineeid).last()
            if nominee:
                Database_Filters['nomination__id'] = nominee.id
        else:
            nominee = ''
        
        allpayments = Payments.objects.filter(**Database_Filters).order_by('-id')
        
        paginator = Paginator(allpayments, 100) # Show 100 contacts per page		
        page = request.GET.get('page')
        payments = paginator.get_page(page)

        totalgrossamount_chk = Payments.objects.filter(**Database_Filters).aggregate(Sum('grossamount'))
        totalgrossamount = str(float(totalgrossamount_chk['grossamount__sum'] or 0) or '0')

        totaltransactionamount_chk = Payments.objects.filter(**Database_Filters).aggregate(Sum('transactionamount'))
        totaltransactionamount = str(float(totaltransactionamount_chk['transactionamount__sum'] or 0) or '0')
        
        totalnetamount_chk = Payments.objects.filter(**Database_Filters).aggregate(Sum('netamount'))
        totalnetamount = str(float(totalnetamount_chk['netamount__sum'] or 0) or '0')

        return render(request, self.template_accountpanel, { 'accountinfo':profile, 'schemes':schemes, 'scheme':scheme, 'categories':categories, 'category':category, 'nominees':nominees, 'nominee':nominee, 'payments':payments, 'totalgrossamount':totalgrossamount, 'totaltransactionamount':totaltransactionamount, 'totalnetamount':totalnetamount  })
    
    """This is responsible for detecting post method and processing information """
    def post(self, request):
        return HttpResponseRedirect(build_url('superadmin_payments', kwargs={}, params={}))
        
class AccountPaymentsExcelView(View):
    form_class = ''
    initial = {'': ''}
    template_accountpanel = 'accountpanel.html'

    """This is responsible for detecting get method and redirecting """
    def get(self, request):
        profile = Accounts.objects.filter(user=request.user).last()        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Votes Reports"
            ws.subject = "Votes Reports"
            ws.description = "This is a spreadsheet of " + "Votes Reports"
            ws['A1'] = "Humanitarian Awards Global"
            ws.merge_cells('A1:H1')
            ws['A2'] = 'HAG'
            ws.merge_cells('A2:H2')
            ws['A4'] = "Votes Reports"


            ws['A7'] = 'Category'				
            ws['B7'] = 'Reference No'				
            ws['C7'] = 'Nominee Firstname'
            ws['D7'] = 'Nominee Middlename'	
            ws['E7'] = 'Nominee Lastname'	
            ws['F7'] = 'Nominee Phonenumber'	
            ws['G7'] = 'Nominee Email Address'	
            ws['H7'] = 'Gender'	
            ws['I7'] = 'Age Group'	
            ws['J7'] = 'Nominee Bio'
            ws['K7'] = 'Why Nominee Deserves It'	
            ws['L7'] = 'Efforts by Nominee'	
            ws['M7'] = 'Turning Point by Nominee'	
            ws['N7'] = 'What makes Nominee Exceptional'	
            ws['O7'] = 'Nominee Reference Links'	
            ws['P7'] = 'Nominee Town or City'
            ws['Q7'] = 'Nominee Region'	
            ws['R7'] = 'Nominee Country'
            ws['S7'] = 'Nominee Code'
            ws['T7'] = 'Company Name'
            ws['U7'] = 'Company Industry'
            ws['V7'] = 'Company Bio'
            ws['W7'] = 'Nominator Fullname'
            ws['X7'] = 'Nominator Phonenumber'
            ws['Y7'] = 'Nominator Email Address'
            ws['Z7'] = 'Nominator Bio'
            ws['Z7'] = 'Time Known Nominee'
            ws['Z7'] = 'Relationship with Nominee'

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 70
            ws.column_dimensions['K'].width = 70
            ws.column_dimensions['L'].width = 70
            ws.column_dimensions['M'].width = 70
            ws.column_dimensions['N'].width = 70
            ws.column_dimensions['O'].width = 20
            ws.column_dimensions['P'].width = 20
            ws.column_dimensions['Q'].width = 20
            ws.column_dimensions['R'].width = 20
            ws.column_dimensions['S'].width = 20
            ws.column_dimensions['T'].width = 30
            ws.column_dimensions['U'].width = 20
            ws.column_dimensions['V'].width = 20
            ws.column_dimensions['W'].width = 20
            ws.column_dimensions['X'].width = 20
            ws.column_dimensions['Y'].width = 20
            ws.column_dimensions['Z'].width = 20
            ws.column_dimensions['AA'].width = 20
            ws.column_dimensions['AB'].width = 20

            nominations = Votes.objects.filter().order_by('id').values_list('id', 'category__year', 'category__name', 'referencenumber', 'nominee_firstname', 'nominee_middlename', 'nominee_lastname', 'nominee_phonenumber', 'nominee_emailaddress', 'nominee_gender', 'nominee_agegroup', 'nominee_bio', 'nominee_deserve_info', 'nominee_effort_info', 'nominee_turning_point', 'nominee_exception', 'nominee_reference_links', 'nominee_towncity', 'nominee_region', 'nominee_country', 'nominee_code', 'company_name', 'company_industry', 'company_bio', 'nominator_fullname', 'nominator_phonenumber', 'nominator_emailaddress', 'nominator_bio', 'nominator_knownnominee_time', 'nominator_nominee_relationship')
            ite = 1
            for nomination in nominations:
                ws['A'+str(ite + 7)] = str(nomination[1] or '')
                ws['B'+str(ite + 7)] = str(nomination[2] or '')
                ws['C'+str(ite + 7)] = str(nomination[3] or '')
                ws['D'+str(ite + 7)] = str(nomination[4] or '')
                ws['E'+str(ite + 7)] = str(nomination[5] or '')
                ws['F'+str(ite + 7)] = str(nomination[6] or '')
                ws['G'+str(ite + 7)] = str(nomination[7] or '')
                ws['H'+str(ite + 7)] = str(nomination[8] or '')
                ws['I'+str(ite + 7)] = str(nomination[9] or '')
                ws['J'+str(ite + 7)] = str(nomination[10] or '')
                ws['K'+str(ite + 7)] = str(nomination[11] or '')
                ws['L'+str(ite + 7)] = str(nomination[12] or '')
                ws['M'+str(ite + 7)] = str(nomination[13] or '')
                ws['N'+str(ite + 7)] = str(nomination[14] or '')
                ws['O'+str(ite + 7)] = str(nomination[15] or '')
                ws['P'+str(ite + 7)] = str(nomination[16] or '')
                ws['Q'+str(ite + 7)] = str(nomination[17] or '')
                ws['R'+str(ite + 7)] = str(nomination[18] or '')
                ws['S'+str(ite + 7)] = str(nomination[19] or '')
                ws['T'+str(ite + 7)] = str(nomination[20] or '')
                ws['U'+str(ite + 7)] = str(nomination[21] or '')
                ws['V'+str(ite + 7)] = str(nomination[22] or '')
                ws['W'+str(ite + 7)] = str(nomination[23] or '')
                ws['X'+str(ite + 7)] = str(nomination[24] or '')
                ws['Y'+str(ite + 7)] = str(nomination[25] or '')
                ws['Z'+str(ite + 7)] = str(nomination[26] or '')
                ws['AA'+str(ite + 7)] = str(nomination[27] or '')
                ws['AB'+str(ite + 7)] = str(nomination[28] or '')

                ite += 1
      
            # useractivity = UserActivity(school=profile.school, activity=profile.fullname + ' generated excel reports for ' + str(reporttype or '') +' at ' +str(datetime.now()))
            # useractivity.save()

            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
            ws.page_setup.fitToHeight = 0
            ws.page_setup.fitToWidth = 1
            output = HttpResponse(content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            file_name = "Votes_Reports.xlsx"
            output['Content-Disposition'] = 'attachment; filename='+ file_name
            wb.save(output)
            return output
            
        except Exception as e:
            return HttpResponse(str(e))

        return render(request, self.template_accountpanel, { 
                                                        'accountinfo': profile })

    def post(self, request):
        profile = UserProfile.objects.filter(user=request.user).last()
        if not profile.module_reports:
            return HttpResponseRedirect(reverse('account'))
            
        return render(request, self.template_accountpanel, { 
                                                        'accountinfo': profile })

class BlogView(View):
    template_accountpanel = 'accountpanel.html'
        
    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        page = request.GET.get('page')
        allblog = Blog.objects.filter().order_by('-id')
        paginator = Paginator(allblog, 20)  # Show 20 blog per page
        blogs = paginator.get_page(page)
        return render(request, self.template_accountpanel, { 'blogs':blogs })
    
    """This is responsible for detecting post method and processing information """
    def post(self, request):
        return HttpResponseRedirect(build_url('superadmin_blog', kwargs={  }, params={ }))

class BlogDetailView(View):
    form_class = BlogForm
    template_accountpanel = 'accountpanel.html'
        
    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        identifier = request.GET.get('identifier')
        blog = None
        related_blog = None
        try:
            blog = Blog.objects.get(identifier=identifier)
            related_blog = Blog.objects.filter().exclude(identifier=blog.identifier).order_by('-id')[:10]
        except Blog.DoesNotExist:
            pass
                 
        return render(request, self.template_accountpanel, { 'blog':blog, 'related_blog':related_blog })
    
    
    """This is responsible for detecting post method and processing information """
    def post(self, request):
        identifier = request.POST.get('identifier')
        form = None
        try:
            if identifier:
                blog = Blog.objects.get(identifier=identifier)
                form = self.form_class(request.POST, request.FILES, instance=blog)
                if form.is_valid():
                    # save information
                    instance = form.save(commit=False) 
                    instance.save()
                    messages.success(request, 'You have successfully updated Blog' )
                    return HttpResponseRedirect(build_url('superadmin_blog_detail', kwargs={  }, params={'identifier':instance.identifier }))
            else:
                form = self.form_class(request.POST, request.FILES)  
                if form.is_valid():
                    # save information
                    instance = form.save(commit=False) 
                    instance.save()
                    messages.success(request, 'You have successfully created Blog' )

                    # Post Email
                    post_email = request.POST.get('post_email')
                    if post_email:
                        # # send mail
                        # recipients = str(request.POST.get('subscriberemail') or '')
                        # subject = 'Newsletter Subscription Successful'
                        # listobj = []
                        # sender_message_data = {'subject': subject, 'fromAddr': '', 'fromName': '', 'toAddr': str(recipients or ''),
                        #                        'replyTo': '', 'template': 'emailsubscription', 'rawMessage': '', 'userId': ''}
                        # listobj.append(sender_message_data)

                        # email_result = bulkemailsender(request, listobj)

                        pass

                    return HttpResponseRedirect(build_url('superadmin_blog_detail', kwargs={  }, params={'identifier':instance.identifier }))
                else:
                    messages.error(request, f'Form is invalid due to {form.errors}' )

        except Exception as e:
            messages.error(request, f'Error caused by: {e}')
            
        
       

        return HttpResponseRedirect(build_url('superadmin_blog', kwargs={  }, params={ }))


class AccountMessagingView(View):
    template_accountpanel = 'accountpanel.html'
        
    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        profile = Accounts.objects.filter(user=request.user).last()
        return HttpResponseRedirect(reverse('superadmin_messaging_sms'))
    
    """This is responsible for detecting post method and processing information """
    def post(self, request):
        return HttpResponseRedirect(reverse('superadmin_messaging'))

class AccountMessagingSMSView(View):
    template_accountpanel = 'accountpanel.html'
        
    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        profile = Accounts.objects.filter(user=request.user).last()
        scheme_get = request.GET.get('scheme')
        category_get = request.GET.get('category')
        Database_Filters = {  }

        schemes = AwardSchemes.objects.filter().order_by('year')
        scheme = ''
        if scheme_get:
            scheme = AwardSchemes.objects.filter(id=scheme_get).last()
            Database_Filters['category__scheme'] = scheme
        else:
            scheme = AwardSchemes.objects.filter().last()
            Database_Filters['category__scheme'] = scheme

        categories = NominationCategories.objects.filter().order_by('id')
        category = ''
        if category_get == 'Individually':
            category = 'Individually'
        elif not category == None:
            category = NominationCategories.objects.filter(scheme=scheme, id=category_get).last()
            Database_Filters['category'] = category

        return render(request, self.template_accountpanel, { 'accountinfo':profile, 'schemes':schemes, 'scheme':scheme, 'categories':categories, 'category':category, })
    
    """This is responsible for detecting post method and processing information """
    def post(self, request):
        return HttpResponseRedirect(reverse('superadmin_messaging'))

class AccountMessagingEmailsView(View):
    template_accountpanel = 'accountpanel.html'
        
    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        profile = Accounts.objects.filter(user=request.user).last()
        scheme_get = request.GET.get('scheme')
        category_get = request.GET.get('category')
        Database_Filters = {  }

        schemes = AwardSchemes.objects.filter().order_by('year')
        scheme = ''
        if scheme_get:
            scheme = AwardSchemes.objects.filter(id=scheme_get).last()
            Database_Filters['category__scheme'] = scheme
        else:
            scheme = AwardSchemes.objects.filter().last()
            Database_Filters['category__scheme'] = scheme

        categories = NominationCategories.objects.filter().order_by('id')
        category = ''
        if category_get == 'Individually':
            category = 'Individually'
        elif not category == None:
            category = NominationCategories.objects.filter(scheme=scheme, id=category_get).last()
            Database_Filters['category'] = category

        return render(request, self.template_accountpanel, { 'accountinfo':profile, 'schemes':schemes, 'scheme':scheme, 'categories':categories, 'category':category, })
    
    """This is responsible for detecting post method and processing information """
    def post(self, request):
        return HttpResponseRedirect(reverse('superadmin_messaging'))

class AccountMessagingSettingsView(View):
    template_accountpanel = 'accountpanel.html'
        
    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        profile = Accounts.objects.filter(user=request.user).last()
        return render(request, self.template_accountpanel, { 'accountinfo':profile })
    
    """This is responsible for detecting post method and processing information """
    def post(self, request):
        return HttpResponseRedirect(reverse('superadmin_messaging'))

class AccountProfileView(View):
    form_class = ProfileForm
    template_profile = 'accountpanel.html'
        
    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        profile = Accounts.objects.filter(user=request.user).last()
        return render(request, self.template_profile, { 'accountinfo':profile    })
    
    """This is responsible for detecting post method and processing information """
    def post(self, request):
        profile = Accounts.objects.filter(user=request.user).last()
        form = self.form_class(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            # save information
            instance = form.save(commit=False)	                    

            instance.save()
            messages.success(request, 'You have successfully saved your Profile' )

        # return HttpResponseRedirect(build_url('projecteditoverview', kwargs={'projectid':project.id}, params={	}))

        return HttpResponseRedirect(reverse('superadmin_profile'))

class RegisterView(View):
    form_class = RegisterForm
    initial = {'': ''}
    template_register = 'register.html'
    template_registerthanks = 'register_thanks.html'

    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        logout(request)		
        form = self.form_class(initial=self.initial)
        return render(request, self.template_register, {'form': form,
                                                        })
    
    """This is responsible for detecting post method and processing information """
    def post(self, request):
        # return HttpResponseRedirect(build_url('register', kwargs={}, params={ }))
        logout(request)	
        form = self.form_class(request.POST)
        
        """This checks if form is valid then authenticates and redirects to main page else redirects to login page """
        if form.is_valid():
            # recaptcha_token = self.request.POST.get('g-recaptcha-response')
            # recaptcha_response = verifyrecaptcha(recaptcha_token)
            
            # if not recaptcha_response:
            #     messages.error(request, 'Invalid reCAPTCHA. Please try again.')
            #     return HttpResponseRedirect(build_url('register', kwargs={}, params={ }))                

            # send mail
            emailaddress = form.cleaned_data['emailaddress']
            regcheck = Register.objects.filter(emailaddress=emailaddress).last()
            usercheck = User.objects.filter(email=emailaddress).last()
            if regcheck:
                messages.error(request, 'Account with this email already exists, kindly Log In')
                return HttpResponseRedirect(reverse('login'))
            
            else:
                code = generate_token()	
                reg = Register(fullname=str(request.POST['fullname']), emailaddress=emailaddress, code=code)
                reg.save()
                
                # send mail
                subject = 'User Registration - HAG'
                # will be used when confirm resgistration
                    
                message = 'Welcome to HAG, We are pleased to have you, we hope you will enjoy the experience. One more simple STEP before you start....<br> <a href="https://www.girif.org/register_confirm?emailaddress='+emailaddress+'&code='+code+'" style="margin:20px 0; display:inline-block; padding:20px; border-radius:4px; background-color:#1ABC9C;  color:#fff; text-decoration:none; ">CLICK HERE TO CONFIRM</a><br>. <h4 style="margin:0">Obipiseibima Priscillia Aggokabo</h4> <p style="margin:0">priscillia@girif.org <br> President</p> '
                html_message = placemessageontemplate('', message)
                # send confirmation email
                send_mail(subject, message, settings.EMAIL_ADMIN, [emailaddress], html_message=html_message)
                # save register information									

                return render(request, self.template_registerthanks, {   })            
                        
        else:
            messages.error(request, form.errors)
            messages.error(request, 'Registration Failed')
            return render(request, self.template_register, {    'form' : form, 
                                                                'message': form.errors, 	
                                                            })

class RegisterConfirmView(View):
    form_class = RegisterConfirmForm
    initial = {'': ''}
    template_registerconfirm = 'register_confirm.html'

    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        logout(request)		
        try:
            reg = Register.objects.filter(emailaddress=request.GET.get('emailaddress'), code=request.GET.get('code')).last()
            if reg:
                return render(request, self.template_registerconfirm, {'email': request.GET.get('emailaddress'), 'code': request.GET.get('code')})
            else:
                messages.error(request, 'Registered Account could not be found')
                return HttpResponseRedirect(reverse('register'))
        except Exception as e:
            messages.error(request, str(e))
            return HttpResponseRedirect(reverse('register'))

    """This is responsible for detecting post method and processing information """
    def post(self, request):
        logout(request)	
        form = self.form_class(request.POST)
        if form.is_valid():
            reg = Register.objects.filter(emailaddress=request.POST['emailaddress'], code=request.POST['code']).last()
            if reg:
                user = User.objects.create_user(username=request.POST['username'], email=request.POST['emailaddress'], password=request.POST['confirmpassword'] )
                acc = Accounts(user=user, fullname=reg.fullname, email=reg.emailaddress, code=reg.code )
                acc.save()

                Register.objects.filter(emailaddress=request.POST['emailaddress'], code=request.POST['code']).update(confirmed=True)

                useractivity = UserActivity(user=acc, activity=reg.fullname + ' Account Created by '+ str(reg.fullname) +' at ' +str(datetime.now()))
                useractivity.save()

                # send mail
                recipients = reg.emailaddress
                subject = 'Registration Complete - '+ reg.fullname
                # will be used when confirm resgistration
                    
                message = 'Congratulations!!!. Welcome to HAG, You have successfully completed the registration process. Thank You. <br>. <h4 style="margin:0">Obipiseibima Priscillia Aggokabo</h4> <p style="margin:0">priscillia@girif.org <br> President</p> '
                html_message = placemessageontemplate('', message)

                # send confirmation email
                send_mail(subject, message, settings.EMAIL_ADMIN, [reg.emailaddress], html_message=html_message)
                # save register information	

                user = authenticate(username=request.POST['username'], password=request.POST['confirmpassword'])
                if user:
                    login(request, user)
                    
                    return HttpResponseRedirect(reverse('account'))
            else:
                messages.error(request, 'Account cannot be found')
                return HttpResponseRedirect(reverse('register'))
        else:
            messages.error(request, 'Account could not be created ' + str(form.errors))
            return HttpResponseRedirect(reverse('register'))
        

        return HttpResponseRedirect(reverse('register'))

def csrf_failure_function(request):
    logout(request)	
    return HttpResponseRedirect(build_url('index', kwargs={}, params={}))


def csrf_failure_function(request, reason):
    logout(request)
    return HttpResponseRedirect(build_url('index', kwargs={}, params={}))


class LoginView(View):
    form_class = LoginForm
    initial = {'': ''}
    template_adminlogin = 'adminlogin.html'

    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        logout(request)		
        form = self.form_class(initial=self.initial)
        return render(request, self.template_adminlogin, {'form': form, 'message':'HAG Login'})
    
    """This is responsible for detecting post method and processing information """
    def post(self, request):
        # print('reached')
        logout(request)	
        form = self.form_class(request.POST)

        """This checks if form is valid then authenticates and redirects to main page else redirects to login page """
        if form.is_valid():
            # recaptcha_token = self.request.POST.get('g-recaptcha-response')
            # recaptcha_response = verifyrecaptcha(recaptcha_token)
            
            # if not recaptcha_response:
            #     messages.error(request, 'Invalid reCAPTCHA. Please try again.')
            #     return HttpResponseRedirect(build_url('login', kwargs={}, params={ }))

            user = authenticate(username=request.POST['username'], password=request.POST['password'])
            if user:
                login(request, user)
                # print('logged in')

                # profile = Accounts.objects.filter(user=user).last()
                # useractivity = UserActivity(user=profile, activity=profile.fullname + ' Logged in at ' +str(datetime.now()))
                # useractivity.save()

                next_url = request.POST.get('next')
                if next_url:
                    return HttpResponseRedirect(next_url)
                else:
                    return HttpResponseRedirect(reverse('superadmin_categories'))	
                
                return HttpResponseRedirect(reverse('superadmin_categories'))	
            else:
                messages.error(request, 'Login Failed' )
                return render(request, self.template_adminlogin, {'form' : form	})
        else:
            return render(request, self.template_adminlogin, {'form' : form	})

class ResetPasswordView(View):	
    form_class = ResetPasswordForm
    initial = {'': ''}
    template_resetpassword = 'resetpassword.html'
    template_resetpassword_thanks = 'resetpassword_thanks.html'

    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        logout(request)		
        # form = self.form_class(initial=self.initial)
        return render(request, self.template_resetpassword, {})
    
    """This is responsible for detecting post method and processing information """
    def post(self, request):
        # recaptcha_token = self.request.POST.get('g-recaptcha-response')
        # recaptcha_response = verifyrecaptcha(recaptcha_token)
        
        # if not recaptcha_response:
        #     messages.error(request, 'Invalid reCAPTCHA. Please try again.')
        #     return HttpResponseRedirect(build_url('resetpassword', kwargs={}, params={ }))
        

        profile = Accounts.objects.filter(emailaddress=request.POST['email']).last()
        useractivity = UserActivity(user=profile, activity=profile.fullname + ' made a Reset Password request at ' +str(datetime.now()))
        useractivity.save()

        logout(request)	
        form = self.form_class(request.POST)
            
        if form.is_valid():
            recipient = form.cleaned_data['email']
            reg = Register.objects.filter(emailaddress=form.cleaned_data['email']).last()
            if reg:
                subject = 'Password Reset'				
                link = '<a href="https://www.girif.org/resetpassword_confirm?emailaddress=" + recipient + "&code=" + reg.code + "> Click here to Reset Password</a>'
                # html_message = '<div style="border:1px solid #283941; border-top:5px solid #1ABC9C; text-align:center; overflow:hidden; font-family:arial; width:98%; margin:0 auto; font-size:90%;"> <div id="head" style="background:#283941; color:white; width:100%; height: 70px; padding:5px" align="center"> <img src="http://' + request.META["HTTP_HOST"] + '/static/images/logo.png" style="height:70px; width:auto; padding: 2px; margin:0 auto;">  </div> <div id="body" style="text-align:left; padding:10px;"> <h2 style="text-align:center;"><b>Password Reset</b></h2><p><b>Welcome to HAG,</b></p> <p>Please click here to reset your password: ' + link + '.</p> </div> <div id="foot" style="background:#283941; color:white; width:100%; padding:5px; font-size:80%;"> <span style="color:white;">Powered by HAG</span> </div> </div>'
                message = 'Please click here to reset your password: <br>' + link
                html_message = placemessageontemplate('', message)
                
                # send confirmation email
                send_mail(subject, message, settings.EMAIL_ADMIN, [recipient], html_message=html_message)
                # save register information
                    
                return render(request, self.template_resetpassword_thanks, {})
            else:
                messages.error(request, 'Registered Account could not be found')
                return HttpResponseRedirect(reverse('resetpassword'))	

        else:
            messages.error(request, form.errors)
            messages.error(request, 'Password Reset Failed')
            return HttpResponseRedirect(reverse('resetpassword'))

class ResetPasswordConfirmView(View):
    form_class = ResetPasswordConfirmForm
    initial = {'': ''}
    template_resetpasswordconfirm = 'resetpassword_confirm.html'

    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        # logout(request)		
        try:
            reg = Register.objects.filter(emailaddress=request.GET.get('emailaddress'), code=request.GET.get('code')).last()
            if reg:
                return render(request, self.template_resetpasswordconfirm, {'email': request.GET.get('emailaddress'), 'code': request.GET.get('code')})
            else:
                messages.error(request, 'Registered Account could not be found')
                return HttpResponseRedirect(reverse('register'))
        except Exception as e:
            messages.error(request, str(e))
            return HttpResponseRedirect(reverse('resetpassword'))

    """This is responsible for detecting post method and processing information """
    def post(self, request):
        logout(request)	
        form = self.form_class(request.POST)
        if form.is_valid():
            # recaptcha_token = self.request.POST.get('g-recaptcha-response')
            # recaptcha_response = verifyrecaptcha(recaptcha_token)
            
            # if not recaptcha_response:
            #     messages.error(request, 'Invalid reCAPTCHA. Please try again.')
            #     return HttpResponseRedirect(build_url('resetpassword', kwargs={}, params={ }))

            reg = Register.objects.filter(emailaddress=request.POST['emailaddress'], code=request.POST['code']).last()
            if reg:
                getuser = User.objects.filter(email=request.POST['emailaddress'])								.last()
                getuser.set_password(form.cleaned_data['confirmpassword'])
                getuser.save()

                profile = Accounts.objects.filter(user=getuser).last()
                useractivity = UserActivity(user=profile, activity=profile.fullname + ' successfully resetted Password at ' +str(datetime.now()))
                useractivity.save()

                messages.success(request, 'Password Reset Successful' )
                return HttpResponseRedirect(build_url('adminlogin', kwargs={}, params={ }))
            else:
                messages.error(request, 'Account cannot be found')
                return HttpResponseRedirect(build_url('register', kwargs={}, params={ }))
        else:
            messages.error(request, form.errors)
            messages.error(request, 'Account could not be created')
            return HttpResponseRedirect(build_url('register', kwargs={}, params={ }))
        

        return HttpResponseRedirect(build_url('register', kwargs={}, params={ }))

class LogoutView(View):
    """This is responsible for detecting get method and redirecting """		
    def get(self, request):
        profile = Accounts.objects.filter(user=request.user).last()
        if profile:  
            useractivity = UserActivity(user=profile, activity=profile.fullname + ' Logged out at ' +str(datetime.now()))
            useractivity.save()

        logout(request)		
        return HttpResponseRedirect(reverse('adminlogin'))
    
    """This is responsible for detecting post method and processing information """
    def post(self, request):
        logout(request)	
        return HttpResponseRedirect(reverse('adminlogin'))





